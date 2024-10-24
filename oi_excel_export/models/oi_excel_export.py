"""
Created on Oct 20, 2016

@author: Zuhair
"""

from odoo import models, fields, api, _
import tempfile
import xlsxwriter
import openpyxl
from xlsxwriter.utility import supported_datetime, xl_col_to_name
import os
import logging
from odoo.tools import config

# from xlsxwriter.compatibility import num_types, str_types
from subprocess import call
from odoo.tools.misc import find_in_path
import sys
import shutil
from collections import OrderedDict
import base64
import io
import json
from pytz import timezone
from .copier import MyWorksheetCopy
from copy import copy
from odoo.tools.pdf import merge_pdf
import re
from openpyxl.worksheet.cell_range import CellRange
from odoo.tools.safe_eval import safe_eval
from collections import defaultdict
from xlsxwriter.worksheet import cell_number_tuple
from openpyxl.workbook.protection import WorkbookProtection
from odoo.modules.module import get_module_path

from odoo.exceptions import UserError


try:
    import jpype  # @UnusedImport
    import jpype.imports  # @UnusedImport
except:
    jpype = None
soffice_bin = None
if sys.platform == "win32":
    try:
        soffice_bin = find_in_path("soffice.exe")
    except:
        pass
else:
    try:
        soffice_bin = find_in_path("soffice")
    except:
        pass

_logger = logging.getLogger(__name__)
_debug = config.get("debug.oi_excel_export")

if not soffice_bin:
    _logger.warning("cannot find soffice bin")


class Report(models.TransientModel):
    _name = "oi_excel_export"
    _description = "oi_excel_export"

    datas = fields.Binary(attachment=False)
    filename = fields.Char()
    pdf = fields.Boolean()
    mimetype = fields.Char()

    def _remove_unsupported(self, rows):
        if not rows:
            return []

        def object_info(value):
            return "[%s] %s" % (type(value), getattr(value, "__len__", lambda: "")())

        def type_check(value):
            if value is None:
                return value
            if isinstance(value, bool):
                return value
            if supported_datetime(value):
                user_tz = self.env.user.tz or "UTC"
                local = timezone(user_tz)
                date_on = (
                    fields.Datetime.from_string(value)
                    .replace(tzinfo=timezone("UTC"))
                    .astimezone(local)
                ).replace(tzinfo=None)
                value = date_on
                return value

            if isinstance(value, dict):
                return value.get(self.env.lang) or value.get("en_US") or repr(value)
            try:
                float(value)
                return value
            except ValueError:
                pass
            except TypeError:
                return object_info(value)
            try:
                str(value)
                return value
            except ValueError:
                return object_info(value)

        new_rows = []
        for row in rows:
            new_row = []
            for value in row:
                new_row.append(type_check(value))
            new_rows.append(new_row)
        return new_rows

    @api.model
    def render_qweb_report(self, report, docids, data=None, filename="report.xlsx"):
        content = report._render_qweb_html(docids, data=data)[0]
        tempdir = tempfile.mkdtemp()
        html_file = os.path.join(tempdir, "file.html")
        xlsx_file = os.path.join(tempdir, "file.xlsx")

        with open(html_file, "wb") as f:
            f.write(content)

        call(
            [
                soffice_bin,
                "--headless",
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                tempdir,
                html_file,
            ]
        )

        with open(xlsx_file, "rb") as f:
            datas = base64.b64encode(f.read())

        record = self.sudo().create(
            {"filename": filename, "pdf": False, "datas": datas}
        )
        self.env.cr.execute(
            "update %s set create_uid = %s" % (self._table, self.env.user.id)
        )
        record.invalidate_cache(["create_uid"])

        return record

    @api.model
    def render_qweb_report_odt(self, report, docids, data=None, filename="report.odt"):
        content = report._render_qweb_html(docids, data=data)[0]
        tempdir = tempfile.mkdtemp()
        html_file = os.path.join(tempdir, "file.html")
        odt_file = os.path.join(tempdir, "file.odt")

        with open(html_file, "wb") as f:
            f.write(content)

        call(
            [
                soffice_bin,
                "--headless",
                "--convert-to",
                "odt",
                "--outdir",
                tempdir,
                html_file,
            ]
        )

        with open(odt_file, "rb") as f:
            datas = base64.b64encode(f.read())

        record = self.sudo().create(
            {"filename": filename, "pdf": False, "datas": datas}
        )
        self.env.cr.execute(
            "update %s set create_uid = %s" % (self._table, self.env.user.id)
        )
        record.invalidate_cache(["create_uid"])

        return record

    @api.model
    def render_qweb_report_docx(
        self, report, docids, data=None, filename="report.docx"
    ):
        content = report._render_qweb_html(docids, data=data)[0]
        tempdir = tempfile.mkdtemp()
        html_file = os.path.join(tempdir, "file.html")
        odt_file = os.path.join(tempdir, "file.odt")
        docx_file = os.path.join(tempdir, "file.docx")

        with open(html_file, "wb") as f:
            f.write(content)

        call(
            [
                soffice_bin,
                "--headless",
                "--convert-to",
                "odt",
                "--outdir",
                tempdir,
                html_file,
            ]
        )
        call(
            [
                soffice_bin,
                "--headless",
                "--convert-to",
                "docx",
                "--outdir",
                tempdir,
                odt_file,
            ]
        )

        with open(docx_file, "rb") as f:
            datas = base64.b64encode(f.read())

        record = self.sudo().create(
            {"filename": filename, "pdf": False, "datas": datas}
        )
        self.env.cr.execute(
            "update %s set create_uid = %s" % (self._table, self.env.user.id)
        )
        record.invalidate_cache(["create_uid"])

        return record

    def _action_protect(self, password):
        self.ensure_one()
        if self.pdf:
            return self

        content = base64.decodebytes(self.datas)
        buffer = io.BytesIO(content)

        workbook = openpyxl.load_workbook(buffer)
        workbook.security = WorkbookProtection(
            workbookPassword=password,
            revisionsPassword=password,
            lockStructure=True,
            lockWindows=True,
            lockRevision=True,
        )

        for ws in workbook.worksheets:
            ws.protection.sheet = True
            ws.protection.password = password
            ws.protection.enable()

        buffer = io.BytesIO()
        workbook.save(buffer)

        content = buffer.getvalue()
        self.datas = base64.encodebytes(content)

        return self

    def _action_encrypt(self, password):
        self.ensure_one()
        if self.pdf:
            return self

        if not jpype:
            raise UserError(
                "Python module JPype not installed \n please install using (pip install JPype1)"
            )

        content = base64.decodebytes(self.datas)

        tempdir = tempfile.mkdtemp()

        try:

            filename = os.path.join(tempdir, "file.xlsx")
            with open(filename, "wb") as f:
                f.write(content)

            if not jpype.isJVMStarted():
                jvmpath = (
                    self.env["ir.config_parameter"]
                    .sudo()
                    .get_param("oi_excel_export.jvmpath", default=None)
                )

                classpath = os.path.join(
                    get_module_path("oi_excel_export"), "java", "util-1.0.1.jar"
                )

                jpype.startJVM(jvmpath=jvmpath, classpath=[classpath])

            import openinside  # @UnresolvedImport @UnusedImport

            from openinside.excel.util import MyEncryptor  # @UnresolvedImport

            MyEncryptor().encrypt(filename, password)

            with open(filename, "rb") as f:
                content = f.read()

            self.datas = base64.encodebytes(content)

        finally:
            try:
                shutil.rmtree(tempdir)
            except (OSError, IOError):
                _logger.error("cannot remove dir %s", tempdir)

        return self

    def action_download(self):
        return {
            "type": "ir.actions.client",
            "tag": "file_download",
            "params": {
                "model": self._name,
                "field": "datas",
                "id": self.id,
                "filename": self.filename,
                "filename_field": "filename",
                "download": True,
                "mimetype": self.mimetype,
            },
        }

    def action_preview_pdf(self, title=None):
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "target": "current",
            "view_type": "form",
            "view_mode": "pdf",
            "name": title or self.filename,
        }

    def _eval_rows(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        has_formula = False

        is_formula = lambda value: isinstance(value, str) and value.startswith("=")

        for row_idx, row in enumerate(rows):
            for col_idx, value in enumerate(row):
                ws.cell(row_idx + 1, col_idx + 1, value)
                if is_formula(value):
                    has_formula = True

        if not has_formula:
            return rows

        def get_lst(arg):
            lst = []
            for a in arg:
                a = filter(lambda v: isinstance(v, (int, float)), a)
                lst.extend(a)
            return lst

        def SUM(*arg):
            lst = get_lst(arg)
            return sum(lst)

        def AVERAGE(*arg):
            lst = get_lst(arg)
            if not lst:
                return None
            return sum(lst) / len(lst)

        def MIN(*arg):
            lst = get_lst(arg)
            return min(lst) if lst else None

        def MAX(*arg):
            lst = get_lst(arg)
            return max(lst) if lst else None

        localdict = {"SUM": SUM, "AVERAGE": AVERAGE, "MIN": MIN, "MAX": MAX}

        evaluated = True
        while evaluated:
            evaluated = False
            for row_idx, row in enumerate(rows):
                for col_idx, _ in enumerate(row):
                    cell = ws.cell(row_idx + 1, col_idx + 1)
                    value = cell.value
                    if is_formula(value):
                        has_formula = False
                        for cell_range in re.findall("\w+\d+:\w+\d+", value):
                            cr = CellRange(cell_range)
                            res = []
                            for cr_row in cr.rows:
                                for cr_cell in cr_row:
                                    cr_value = ws.cell(*cr_cell).value
                                    if is_formula(cr_value):
                                        has_formula = True
                                        continue
                                    res.append(cr_value)
                            value = value.replace(cell_range, str(res))
                        if not has_formula:
                            value = value[1:]
                            try:
                                value = safe_eval(value, localdict)
                            except:
                                value = None
                            cell.value = value
                            evaluated = True

        for row_idx, row in enumerate(rows):
            for col_idx, value in enumerate(row):
                cell = ws.cell(row_idx + 1, col_idx + 1)
                if is_formula(value):
                    row[col_idx] = cell.value

        return row

    @api.model
    def _merge_pdf(self, pdf_data, filename="report.pdf"):
        data = merge_pdf(pdf_data)
        datas = base64.encodebytes(data)
        return self.create(
            {
                "filename": filename,
                "datas": datas,
                "pdf": True,
                "mimetype": "application/pdf",
            }
        )

    def merge_pdf(self, filename=None):
        pdf_data = []
        for record in self:
            data = base64.decodebytes(record.datas)
            pdf_data.append(data)
        return self._merge_pdf(pdf_data, filename or self[0].filename)

    def merge(self, title="", update_title=True, pdf=False):
        wb_res = False
        title = title or _("Page")
        index = 0
        for record in self:
            assert not record.pdf
            value = base64.decodebytes(record.datas)
            with io.BytesIO() as buf:
                buf.write(value)
                workbook = openpyxl.load_workbook(buf)
                if not wb_res:
                    wb_res = workbook
                    for ws in workbook.worksheets:
                        index += 1
                        if update_title:
                            ws.title = "%s %d" % (title, index)
                    continue
                else:
                    pass
                for ws in workbook.worksheets:
                    borders = {}
                    for rc in ws._cells:
                        cell = ws.cell(*rc)
                        borders[rc] = copy(cell.border)
                    index += 1
                    if update_title:
                        new_title = "%s %d" % (title, index)
                    else:
                        new_title = ws.title
                    new_ws = wb_res.create_sheet(new_title)
                    cp = MyWorksheetCopy(source_worksheet=ws, target_worksheet=new_ws)
                    cp.copy_worksheet()
                    for name in [
                        "views",
                        "_images",
                        "page_setup",
                        "page_margins",
                        "orientation",
                        "paper_size",
                        "print_options",
                        "firstHeader",
                        "oddHeader",
                        "evenHeader",
                        "firstFooter",
                        "oddFooter",
                        "evenFooter",
                        "sheet_properties",
                        "sheet_format",
                    ]:
                        setattr(new_ws, name, copy(getattr(ws, name)))
                    for rc in new_ws._cells:
                        if rc in borders:
                            cell = new_ws.cell(*rc)
                            cell.border = borders[rc]

        if wb_res:
            filename = self[0].filename
            if pdf and soffice_bin:
                filename = filename[:-4] + "pdf"
                tempdir = tempfile.mkdtemp()
                xlsx_file = os.path.join(tempdir, "file.xlsx")
                wb_res.save(xlsx_file)
                pdf_file = os.path.join(tempdir, "file.pdf")
                call(
                    [
                        soffice_bin,
                        "--headless",
                        "--convert-to",
                        "pdf",
                        "--outdir",
                        tempdir,
                        xlsx_file,
                    ]
                )
                with open(pdf_file, "rb") as f:
                    datas = base64.encodebytes(f.read())
            else:
                with io.BytesIO() as buf:
                    wb_res.save(buf)
                    datas = base64.encodebytes(buf.getvalue())

            return self.create({"filename": filename, "datas": datas})
        return self.browse()

    @api.model
    def export_qweb(self, data, report_name="oi_excel_export.report_oi_excel_export"):
        report = self.env["ir.actions.report"]._get_report_from_name(report_name)
        defaults = {
            "layout": "internal",
            "report_title": report.name,
            "header_rows_count": 1,
        }

        for name, value in defaults.items():
            data.setdefault(name, value)

        row_merge_cells_skip = {}
        row_merge_cells_vals = {}
        if data.get("row_merge_cells"):
            for row, col1, col2 in data["row_merge_cells"]:
                row_merge_cells_vals["%s,%s" % (row, col1)] = col2 - col1 + 1
                for i in range(col1 + 1, col2 + 1):
                    row_merge_cells_skip["%s,%s" % (row, i)] = True

        data["row_merge_cells_vals"] = row_merge_cells_vals
        data["row_merge_cells_skip"] = row_merge_cells_skip

        if "number_columns" not in data:
            number_columns = set()
            for row in data.get("rows"):
                for col, value in enumerate(row):
                    if col in number_columns:
                        continue
                    if isinstance(value, (float, int)):
                        number_columns.add(col)

            data["number_columns"] = list(number_columns)

        record = self.create({"datas": json.dumps(data)})
        return report.report_action(record.ids, data={})

    @api.model
    def export(
        self,
        rows,
        parameter=None,
        decimal_places=2,
        row_color1="white",
        row_color2="#f8f9f9",
        summary_color="#ebedef",
        header_color="#ebedef",
        group_color="#B0D1F7",
        title_color="#B0C4DE",
        date_format="yyyy-MM-dd",
        filename="report",
        add_row_total="SUM",
        add_column_total=None,
        header_rows_count=1,
        header_columns_count=0,
        date_cols=[],
        summary_rows=[],
        group_rows=[],
        total_rows=[],
        total_columns=[],
        percentage_columns=[],
        percentage_rows=[],
        row_merge_cells=[],
        worksheet_name="data",
        pdf=False,
        empty_rows=[],
        title_rows=[],
        action=True,
        outlines_row=[],
        outline_on=True,
        outline_below=True,
        outline_right=True,
        auto_style=False,
        images=[],
        after_func=None,
        openpyxl_func=None,
        freeze_panes=True,
        formats={},
        char_width=1.0,
        min_width=10,
        max_width=100,
        landscape=None,
        html=False,
        font_size=None,
        data_source=None,
        page_header="",
        hidden_columns=[],
        hidden_rows=[],
    ):

        if rows and isinstance(rows[0], OrderedDict):
            res = rows
            rows = []
            rows.append(res[0].keys())
            for vals in res:
                rows.append(vals.values())

        rows = self._remove_unsupported(rows)

        filename = filename or "report"

        # filename = re.sub('[^0-9a-zA-Z]+', '_', filename)
        filename = filename.replace(os.sep, "_")
        filename = filename.replace("/", "_")

        if not filename.endswith(".xlsx"):
            filename = filename + ".xlsx"

        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        column_names = rows[0] if rows else []
        column_counts = len(column_names)
        tempdir = tempfile.mkdtemp()
        temp_filename = os.path.join(tempdir, filename)
        temp_file = open(temp_filename, "wb+")
        workbook = xlsxwriter.Workbook(temp_file)

        right_to_left = (
            self.env["res.lang"]._lang_get(self._context.get("lang")).direction == "rtl"
        )
        reading_order = right_to_left and 2 or 1

        worksheet = workbook.add_worksheet(name=worksheet_name)
        if right_to_left:
            worksheet.right_to_left()
        worksheet.outline_settings(outline_on, outline_below, outline_right, auto_style)
        row_color = (row_color2, row_color1, summary_color, group_color)
        num_format = "#,##0"
        if decimal_places > 0:
            num_format = num_format + "." + "0" * decimal_places
        percent_format = "0.0%"

        num_formats = [
            {
                "bg_color": row_color[0],
                "num_format": num_format,
                "left": 1,
                "right": 1,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[1],
                "num_format": num_format,
                "left": 1,
                "right": 1,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[2],
                "bold": True,
                "num_format": num_format,
                "left": 1,
                "right": 1,
                "top": 2,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[2],
                "bold": True,
                "num_format": num_format,
                "left": 1,
                "right": 1,
                "top": 2,
                "bottom": 2,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[3],
                "bold": True,
                "num_format": num_format,
                "left": 1,
                "right": 1,
                "top": 2,
                "bottom": 2,
                "reading_order": reading_order,
            },
        ]

        for index, vals in enumerate(num_formats):
            vals.update(formats.get("num_formats-%d" % index, {}))

        num_formats = tuple(map(lambda d: workbook.add_format(d), num_formats))

        percent_formats = [
            {
                "bg_color": row_color[0],
                "num_format": percent_format,
                "left": 1,
                "right": 1,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[1],
                "num_format": percent_format,
                "left": 1,
                "right": 1,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[2],
                "bold": True,
                "num_format": percent_format,
                "left": 1,
                "right": 1,
                "top": 2,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[2],
                "bold": True,
                "num_format": percent_format,
                "left": 1,
                "right": 1,
                "top": 2,
                "bottom": 2,
                "reading_order": reading_order,
            },
        ]

        for index, vals in enumerate(percent_formats):
            vals.update(formats.get("percent_formats-%d" % index, {}))

        percent_formats = tuple(map(lambda d: workbook.add_format(d), percent_formats))

        date_formats = [
            {
                "bg_color": row_color[0],
                "num_format": date_format,
                "left": 1,
                "right": 1,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[1],
                "num_format": date_format,
                "left": 1,
                "right": 1,
                "reading_order": reading_order,
            },
        ]

        for index, vals in enumerate(date_formats):
            vals.update(formats.get("date_formats-%d" % index, {}))

        date_formats = tuple(map(lambda d: workbook.add_format(d), date_formats))

        row_format = [
            {
                "bg_color": row_color[0],
                "left": 1,
                "right": 1,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[1],
                "left": 1,
                "right": 1,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[2],
                "bold": True,
                "left": 0,
                "right": 0,
                "top": 2,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[2],
                "bold": True,
                "left": 0,
                "right": 0,
                "top": 2,
                "bottom": 2,
                "reading_order": reading_order,
            },
            {
                "bg_color": row_color[3],
                "bold": True,
                "left": 0,
                "right": 0,
                "top": 2,
                "bottom": 2,
                "reading_order": reading_order,
            },
        ]

        for index, vals in enumerate(row_format):
            vals.update(formats.get("row_format-%d" % index, {}))

        row_format = tuple(map(lambda d: workbook.add_format(d), row_format))

        last_row_format = {
            "left": 0,
            "right": 0,
            "top": 2,
            "reading_order": reading_order,
        }
        last_row_format.update(formats.get("last_row_format", {}))
        last_row_format = workbook.add_format(last_row_format)

        header_format = {
            "bold": True,
            "bg_color": header_color,
            "text_h_align": right_to_left and 3 or 1,
            "text_v_align": 1,
            "bottom_color": "black",
            "top": 2,
            "bottom": 2,
            "left": 1,
            "right": 1,
            "reading_order": reading_order,
        }
        header_format.update(formats.get("header_format", {}))
        header_format = workbook.add_format(header_format)

        header_format2 = {
            "bold": True,
            "bg_color": header_color,
            "text_h_align": right_to_left and 3 or 1,
            "text_v_align": 1,
            "bottom_color": "black",
            "top": 2,
            "bottom": 2,
            "left": 0,
            "right": 0,
            "reading_order": reading_order,
        }
        header_format2.update(formats.get("header_format2", {}))
        header_format2 = workbook.add_format(header_format2)

        empty_format = {
            "bg_color": "#ffffff",
            "text_h_align": right_to_left and 3 or 1,
            "text_v_align": 1,
            "bottom_color": "black",
            "top": 0,
            "bottom": 0,
            "left": 0,
            "right": 0,
            "reading_order": reading_order,
        }
        empty_format.update(formats.get("empty_format", {}))
        empty_format = workbook.add_format(empty_format)

        title_format = {
            "bold": True,
            "bg_color": title_color,
            "text_h_align": right_to_left and 3 or 1,
            "text_v_align": 1,
            "bottom_color": "black",
            "top": 2,
            "bottom": 2,
            "left": 0,
            "right": 0,
            "reading_order": reading_order,
        }
        title_format.update(formats.get("title_format", {}))
        title_format_num = title_format.copy()

        title_format = workbook.add_format(title_format)

        title_format_num.update(
            {"reading_order": reading_order, "num_format": num_format, "align": "right"}
        )

        title_format_num.update(formats.get("title_format_num", {}))
        title_format_num = workbook.add_format(title_format_num)

        if html and font_size is None:
            font_size = 20

        if font_size is not None:
            for workbook_format in workbook.formats:
                workbook_format.set_font_size(font_size)

        def cell_width(value):
            if isinstance(value, float):
                str_format = "{:,.%sf}" % decimal_places
                value = str_format.format(value)
            if not isinstance(value, str):
                value = str(value)
            if value.startswith("="):
                value = "9,999,999.00"
            return max(min(len(value) * char_width + 1, max_width), min_width)

        for row_no in outlines_row:
            worksheet.set_row(row_no, options={"level": 1, "hidden": True})

        if parameter:
            worksheet2 = workbook.add_worksheet(name="parameter")
            if right_to_left:
                worksheet2.right_to_left()
            max_width = []
            for row_no, row in enumerate(parameter):
                for col_no, value in enumerate(row):
                    fn = lambda: header_format if col_no == 0 else None
                    if value:
                        worksheet2.write(row_no, col_no, value, fn())
                    while len(max_width) <= col_no:
                        max_width.append(10)
                    max_width[col_no] = max(max_width[col_no], cell_width(value))

            if _debug:
                _logger.info("parameter max_width " + str(max_width))
            for col_no, col_width in enumerate(max_width):
                worksheet2.set_column(col_no, col_no, col_width)
            worksheet2.set_margins(left=0.25, right=0.25, top=0.5, bottom=0.5)
            worksheet2.set_header(
                "&L%s - parameter&R%s"
                % (
                    page_header or filename.replace(".xlsx", "").replace("_", " "),
                    self.env.company.name,
                )
            )
            worksheet2.set_footer("&L%s&RPage &P of &N" % (fields.Datetime.now()))

        num_columns = set()

        def get_format(row_no, col_no, value):
            try:
                if row_no < header_rows_count:
                    if col_no < header_columns_count:
                        if not value or not str(value).strip():
                            return None
                    return header_format
                if row_no in group_rows:
                    index = 4
                elif row_no in summary_rows:
                    index = 3
                elif row_no in empty_rows:
                    return empty_format
                elif row_no in title_rows:
                    if isinstance(value, float) or (
                        isinstance(value, str) and value.startswith("=")
                    ):
                        return title_format_num
                    return title_format
                elif col_no < header_columns_count:
                    return header_format2
                else:
                    index = row_no % 2
                if isinstance(value, float) or (
                    isinstance(value, str) and value.startswith("=")
                ):
                    if col_no in percentage_columns:
                        return percent_formats[index]
                    if row_no in percentage_rows:
                        return percent_formats[index]
                    return num_formats[index]
                if supported_datetime(value):
                    return date_formats[index]
                return row_format[index]
            except:
                return None

        col_width = []

        def set_col_width(col_no, value):
            while col_no >= len(col_width):
                col_width.append(None)
            col_width[col_no] = max(col_width[col_no] or 0, cell_width(value) or 0)

        def get_date_value(value):
            try:
                if isinstance(value, str):
                    return fields.Date.from_string(value)
            except:
                pass
            return value

        def sum_column(col_no, start, end):
            total = 0.0
            for row in rows:
                if len(row) > col_no:
                    value = row[col_no]
                    if isinstance(value, float):
                        total += value
            return total

        def sum_row(row_no, start, end):
            total = 0.0
            if row_no >= len(rows):
                return total
            row = rows[row_no]
            for col_no in range(start, end + 1):
                if len(row) > col_no:
                    value = row[col_no]
                    if isinstance(value, float):
                        total += value
            return total

        for row_no, row in enumerate(rows):
            if row_no in group_rows:
                first_col = len(row) - 1
                last_col = len(column_names) - 1
            else:
                first_col = None
                last_col = None
            merge_col = []
            for _row, _first_col, _last_col in row_merge_cells:
                if _row == row_no:
                    merge_col.append((_first_col, _last_col))
            for col_no, value in enumerate(row):
                if isinstance(value, float):
                    num_columns.add(col_no)
                if isinstance(value, str) and value.startswith("="):
                    num_columns.add(col_no)
                if col_no in date_cols and row_no > 0:
                    value = get_date_value(value)
                _first_col, _last_col = None, None
                for _first_col, _last_col in merge_col:
                    if _first_col == col_no:
                        break
                if col_no == _first_col and _last_col > _first_col:
                    worksheet.merge_range(
                        row_no,
                        _first_col,
                        row_no,
                        _last_col,
                        value,
                        get_format(row_no, col_no, value),
                    )
                elif col_no == first_col and last_col > first_col:
                    worksheet.merge_range(
                        row_no,
                        first_col,
                        row_no,
                        last_col,
                        value,
                        get_format(row_no, col_no, value),
                    )
                else:
                    set_col_width(col_no, value)
                    worksheet.write(
                        row_no, col_no, value, get_format(row_no, col_no, value)
                    )

        if add_row_total:
            if isinstance(add_row_total, bool):
                add_row_total = "SUM"
            if not total_rows:
                total_rows = [len(rows)]

            total_total = []
            if len(total_rows) > 1:
                for index in range(0, column_counts):  # @UnusedVariable
                    total_total.append([])

            first_row = header_rows_count + 1

            for last_row in total_rows:
                for col_no in range(header_columns_count, len(column_names)):
                    if add_row_total in ["COUNTA"]:
                        col_name = xl_col_to_name(col_no)
                        formola = "=%s(%s%s:%s%s)" % (
                            add_row_total,
                            col_name,
                            first_row,
                            col_name,
                            last_row,
                        )
                        worksheet.write_formula(
                            last_row, col_no, formola, num_formats[2]
                        )
                        if total_total:
                            total_total[col_no].append(
                                "%s%s" % (col_name, last_row + 1)
                            )
                    elif col_no not in num_columns or col_no in percentage_columns:
                        worksheet.write(last_row, col_no, "", row_format[2])

                for col_no in num_columns:
                    if col_no in percentage_columns:
                        continue
                    col_name = xl_col_to_name(col_no)
                    formola = "=%s(%s%s:%s%s)" % (
                        add_row_total,
                        col_name,
                        first_row,
                        col_name,
                        last_row,
                    )
                    set_col_width(col_no, sum_column(col_no, first_row, last_row))
                    worksheet.write_formula(last_row, col_no, formola, num_formats[2])
                    if total_total:
                        total_total[col_no].append("%s%s" % (col_name, last_row + 1))

                first_row = last_row + 3

            last_row = len(rows)

            for col_no, total in enumerate(total_total):
                if total:
                    formola = "=%s(%s)" % (add_row_total, ",".join(total))
                    worksheet.write_formula(last_row, col_no, formola, num_formats[2])
                else:
                    worksheet.write(last_row, col_no, "", row_format[2])
        else:
            last_row = len(rows)
            for col_no in range(header_columns_count, len(column_names)):
                worksheet.write(last_row, col_no, "", last_row_format)

        if add_column_total:
            if isinstance(add_column_total, bool):
                add_column_total = "SUM"
            if not total_columns:
                total_columns = [column_counts]

            total_total = []
            if len(total_columns) > 1:
                for index in range(0, len(rows)):  # @UnusedVariable
                    total_total.append([])

            first_col = header_columns_count

            for last_col in total_columns:

                col_total = 0

                for row_no in range(
                    header_columns_count, len(rows) + (1 if add_row_total else 0)
                ):
                    if row_no in group_rows:
                        worksheet.write(row_no, last_col, "", row_format[4])
                        continue
                    first_col_name = xl_col_to_name(first_col)
                    last_col_name = xl_col_to_name(last_col - 1)
                    formola = "=%s(%s%s:%s%s)" % (
                        add_column_total,
                        first_col_name,
                        row_no + 1,
                        last_col_name,
                        row_no + 1,
                    )
                    worksheet.write_formula(row_no, last_col, formola, num_formats[2])
                    col_total = max(col_total, sum_row(row_no, first_col, last_col))
                    if total_total:
                        total_total[row_no].append("%s%s" % (last_col_name, row_no + 1))

                set_col_width(last_col, col_total)

                first_col = last_col + 3

            last_col = column_counts

            for row_no, total in enumerate(total_total):
                if total:
                    formola = "=%s(%s)" % (add_column_total, ",".join(total))
                    worksheet.write_formula(row_no, last_col, formola, num_formats[2])
                else:
                    worksheet.write(row_no, last_col, "", row_format[2])

        total_width = 0
        for col, width in enumerate(col_width):
            total_width += width
            if width:
                worksheet.set_column(col, col, width)

        for row in hidden_rows:
            worksheet.set_row(row, row, options={"hidden": True})

        for col in hidden_columns:
            worksheet.set_column(col, col, options={"hidden": True})

        worksheet.set_paper(9)
        if total_width > 100 and landscape is None:
            worksheet.set_landscape()
        if landscape:
            worksheet.set_landscape()

        worksheet.set_margins(left=0.25, right=0.25, top=0.5, bottom=0.5)
        worksheet.set_header(
            "&L%s&R%s"
            % (
                filename.replace(".xlsx", "").replace("_", " "),
                self.env.company.name,
            )
        )
        worksheet.set_footer("&L%s&RPage &P of &N" % (fields.Datetime.now()))

        if freeze_panes:
            if header_rows_count or header_columns_count:
                worksheet.freeze_panes(header_rows_count, header_columns_count)

        for image in images:
            if image.get("data"):
                with tempfile.NamedTemporaryFile(
                    "wb+", suffix=image.get("name"), delete=False
                ) as temp:
                    temp.write(base64.b64decode(image["data"]))
                    image["file"] = temp.name

            if image.get("file"):
                worksheet.insert_image(
                    image["cell"], image["file"], image.get("dimensions", {})
                )

        if data_source:
            for row_no, col_no, details_rows, _ in data_source:
                cell_name = "%s%d" % (xl_col_to_name(col_no), row_no + 1)
                cell = worksheet.table[row_no][col_no]
                if not isinstance(cell, cell_number_tuple):
                    continue
                ws = workbook.add_worksheet(name=cell_name)
                ws.set_margins(left=0.25, right=0.25, top=0.75, bottom=0.5)
                ws.set_header(
                    "&L%s-%s(%s)&R%s"
                    % (
                        filename.replace(".xlsx", "").replace("_", " "),
                        cell_name,
                        cell.number,
                        self.env.company.name,
                    )
                )
                ws.set_footer("&L%s&RPage &P of &N" % (fields.Datetime.now()))
                ws.set_landscape()
                ws.set_paper(9)
                ws.fit_to_pages(1, 0)

                worksheet.write_url(
                    row_no,
                    col_no,
                    "internal:%s!A:A" % cell_name,
                    string=str(cell.number),
                    cell_format=cell.format,
                )
                worksheet.write_number(
                    row_no, col_no, cell.number, cell_format=cell.format
                )

                ws.freeze_panes(1, 0)
                if right_to_left:
                    ws.right_to_left()
                col_max_width = defaultdict(lambda: 10)
                for row_idx, row in enumerate(details_rows):
                    index = row_idx % 2
                    for col_idx, value in enumerate(row):
                        if row_idx == 0:
                            cell_format = header_format
                        else:
                            if isinstance(value, float):
                                cell_format = num_formats[index]
                            elif supported_datetime(value):
                                cell_format = date_formats[index]
                            else:
                                cell_format = row_format[index]

                        if row_idx == 0 and col_idx == 0:
                            ws.write_url(
                                row_idx,
                                col_idx,
                                "internal:%s!%s" % (worksheet_name, cell_name),
                                string=value,
                                cell_format=cell_format,
                            )
                        else:
                            ws.write(row_idx, col_idx, value, cell_format)

                        col_max_width[col_idx] = max(
                            col_max_width[col_idx], cell_width(value)
                        )

                for col, width in col_max_width.items():
                    if width:
                        ws.set_column(col, col, width)

        if callable(after_func):
            after_func(workbook=workbook, worksheet=worksheet)

        workbook.close()
        temp_file.flush()
        temp_file.seek(0)

        if callable(openpyxl_func):
            temp_file.close()
            workbook = openpyxl.load_workbook(temp_file.name)
            worksheet = workbook.worksheets[0]

            def _setattr(obj, name, value):
                assert type(obj).__module__[:9] == "openpyxl."
                setattr(obj, name, value)

            openpyxl_func(workbook, worksheet, _setattr)
            workbook.save(temp_file.name)
            temp_file = open(temp_file.name, "rb")

        if not soffice_bin:
            pdf = False

        if pdf:
            temp_file.close()
            call(
                [
                    soffice_bin,
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    tempdir,
                    temp_filename,
                ]
            )
            filename = filename.replace(".xlsx", ".pdf")
            temp_filename = os.path.join(tempdir, filename)
            temp_file = open(temp_filename, "rb")
            mimetype = "application/pdf"
        elif html:
            temp_file.close()
            call(
                [
                    soffice_bin,
                    "--headless",
                    "--convert-to",
                    "html",
                    "--outdir",
                    tempdir,
                    temp_filename,
                ]
            )
            filename = filename.replace(".xlsx", ".html")
            temp_filename = os.path.join(tempdir, filename)
            temp_file = open(temp_filename, "rb")
            mimetype = "text/html"

        if action:
            record = self.sudo().create(
                {
                    "filename": filename,
                    "pdf": pdf,
                    "datas": base64.b64encode(temp_file.read()),
                    "mimetype": mimetype,
                }
            )

            assert record.filename

            temp_file.close()
            try:
                shutil.rmtree(tempdir)
            except (OSError, IOError):
                _logger.error("cannot remove dir %s", tempdir)

            if action == "record":
                return record.ids

            if action == "object":
                return record

            return record.action_download()
        return temp_file.read(), pdf and "pdf" or "xlsx"

    @api.constrains("mimetype", "datas")
    def _check_web_asset_mimetype(self):
        for record in self:
            att = self.env["ir.attachment"].search(
                [
                    ("res_model", "=", self._name),
                    ("res_field", "=", "datas"),
                    ("res_id", "=", record.id),
                ]
            )
            if att:
                if att.mimetype != record.mimetype:
                    att.write({"mimetype": record.mimetype})
